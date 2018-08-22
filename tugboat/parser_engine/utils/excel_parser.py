# Copyright 2018 AT&T Intellectual Property.  All other rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import re
import yaml
from openpyxl import load_workbook

from ..check_exceptions import (
    NoSpecMatched, )


class ExcelParser():
    def __init__(self, file_name, excel_specs):
        self.file_name = file_name
        with open(excel_specs, 'r') as f:
            spec_raw_data = f.read()
        self.excel_specs = yaml.safe_load(spec_raw_data)
        self.wb = load_workbook(file_name, data_only=True)
        self.ipmi_data = {}
        self.hosts = []
        self.spec = None

    @staticmethod
    def sanitize(string):
        return string.replace(' ', '').lower()

    def compare(self, string1, string2):
        return bool(re.search(self.sanitize(string1), self.sanitize(string2)))

    def validate_sheet(self, spec, sheet):
        ws = self.wb[sheet]
        header_row = self.excel_specs['specs'][spec]['header_row']
        ipmi_header = self.excel_specs['specs'][spec]['ipmi_address_header']
        ipmi_column = self.excel_specs['specs'][spec]['ipmi_address_col']
        header_value = ws.cell(row=header_row, column=ipmi_column).value
        return bool(self.compare(ipmi_header, header_value))

    def find_correct_spec(self):
        for spec in self.excel_specs['specs']:
            sheet_name = self.excel_specs['specs'][spec]['ipmi_sheet_name']
            for sheet in self.wb.sheetnames:
                if self.compare(sheet_name, sheet):
                    self.excel_specs['specs'][spec]['ipmi_sheet_name'] = sheet
                    if self.validate_sheet(spec, sheet):
                        return spec
        raise NoSpecMatched(self.excel_specs)

    def get_ipmi_data(self):
        self.spec = self.find_correct_spec()
        sheet_name = self.excel_specs['specs'][self.spec]['ipmi_sheet_name']
        ws = self.wb[sheet_name]
        row = self.excel_specs['specs'][self.spec]['start_row']
        end_row = self.excel_specs['specs'][self.spec]['end_row']
        hostname_col = self.excel_specs['specs'][self.spec]['hostname_col']
        ipmi_address_col = self.excel_specs['specs'][self.spec][
            'ipmi_address_col']
        host_profile_col = self.excel_specs['specs'][self.spec][
            'host_profile_col']
        ipmi_gateway_col = self.excel_specs['specs'][self.spec][
            'ipmi_gateway_col']
        while row <= end_row:
            hostname = self.sanitize(
                ws.cell(row=row, column=hostname_col).value)
            self.hosts.append(hostname)
            ipmi_address = ws.cell(row=row, column=ipmi_address_col).value
            if '/' in ipmi_address:
                ipmi_address = ipmi_address.split('/')[0]
            ipmi_gateway = ws.cell(row=row, column=ipmi_gateway_col).value
            tmp_host_profile = ws.cell(row=row, column=host_profile_col).value
            host_profile = tmp_host_profile.split('-')[1]
            self.ipmi_data[hostname] = {
                'ipmi_address': ipmi_address,
                'ipmi_gateway': ipmi_gateway,
                'host_profile': host_profile,
            }
            row += 1
        return [self.ipmi_data, self.hosts]

    def get_private_vlan_data(self, ws):
        vlan_data = {}
        row = self.excel_specs['specs'][self.spec]['vlan_start_row']
        end_row = self.excel_specs['specs'][self.spec]['vlan_end_row']
        type_col = self.excel_specs['specs'][self.spec]['net_type_col']
        vlan_col = self.excel_specs['specs'][self.spec]['vlan_col']
        while row <= end_row:
            cell_value = ws.cell(row=row, column=type_col).value
            if cell_value:
                vlan = ws.cell(row=row, column=vlan_col).value
                vlan_data[vlan] = cell_value
            row += 1
        return vlan_data

    def get_private_network_data(self):
        sheet_name = self.excel_specs['specs'][self.spec]['private_ip_sheet']
        ws = self.wb[sheet_name]
        vlan_data = self.get_private_vlan_data(ws)
        network_data = {}
        row = self.excel_specs['specs'][self.spec]['net_start_row']
        end_row = self.excel_specs['specs'][self.spec]['net_end_row']
        col = self.excel_specs['specs'][self.spec]['net_col']
        vlan_col = self.excel_specs['specs'][self.spec]['net_vlan_col']
        old_vlan = ''
        while row <= end_row:
            vlan = ws.cell(row=row, column=vlan_col).value
            network = ws.cell(row=row, column=col).value
            if vlan and network:
                net_type = vlan_data[vlan]
                if 'vlan' not in network_data:
                    network_data[net_type] = {
                        'vlan': vlan,
                        'subnet': [],
                    }
            elif not vlan and network:
                vlan = old_vlan
            else:
                row += 1
                continue
            network_data[vlan_data[vlan]]['subnet'].append(network)
            old_vlan = vlan
            row += 1
        for network in network_data:
            if len(network_data[network]['subnet']) > 1:
                network_data[network]['is_common'] = False
            else:
                network_data[network]['is_common'] = True
        return network_data

    def get_public_network_data(self):
        network_data = {}
        sheet_name = self.excel_specs['specs'][self.spec]['public_ip_sheet']
        ws = self.wb[sheet_name]
        oam_row = self.excel_specs['specs'][self.spec]['oam_ip_row']
        oam_col = self.excel_specs['specs'][self.spec]['oam_ip_col']
        oam_vlan_col = self.excel_specs['specs'][self.spec]['oam_vlan_col']
        ingress_row = self.excel_specs['specs'][self.spec]['ingress_ip_row']
        oob_row = self.excel_specs['specs'][self.spec]['oob_net_row']
        col = self.excel_specs['specs'][self.spec]['oob_net_start_col']
        end_col = self.excel_specs['specs'][self.spec]['oob_net_end_col']
        network_data = {
            'oam': {
                'ip': ws.cell(row=oam_row, column=oam_col).value,
                'vlan': ws.cell(row=oam_row, column=oam_vlan_col).value,
            },
            'ingress': ws.cell(row=ingress_row, column=oam_col).value,
        }
        network_data['oob'] = {
            'subnets': [],
        }
        while col <= end_col:
            cell_value = ws.cell(row=oob_row, column=col).value
            if cell_value:
                network_data['oob']['subnets'].append(
                    self.sanitize(cell_value))
            col += 1
        return network_data

    def get_dns_ntp_ldap_data(self):
        dns_ntp_ldap_data = {}
        sheet_name = self.excel_specs['specs'][self.spec]['dns_ntp_ldap_sheet']
        ws = self.wb[sheet_name]
        dns_row = self.excel_specs['specs'][self.spec]['dns_row']
        dns_col = self.excel_specs['specs'][self.spec]['dns_col']
        ntp_row = self.excel_specs['specs'][self.spec]['ntp_row']
        ntp_col = self.excel_specs['specs'][self.spec]['ntp_col']
        domain_row = self.excel_specs['specs'][self.spec]['domain_row']
        domain_col = self.excel_specs['specs'][self.spec]['domain_col']
        ldap_subdomain_row = self.excel_specs['specs'][self.spec][
            'ldap_subdomain_row']
        ldap_col = self.excel_specs['specs'][self.spec]['ldap_col']
        ldap_group_row = self.excel_specs['specs'][self.spec]['ldap_group_row']
        ldap_url_row = self.excel_specs['specs'][self.spec]['ldap_url_row']
        dns_ntp_ldap_data = {
            'dns': ws.cell(row=dns_row, column=dns_col).value,
            'ntp': ws.cell(row=ntp_row, column=ntp_col).value,
            'domain': ws.cell(row=domain_row, column=domain_col).value,
            'ldap': {
                'subdomain': ws.cell(
                    row=ldap_subdomain_row, column=ldap_col).value,
                'common_name': ws.cell(
                    row=ldap_group_row, column=ldap_col).value,
                'url': ws.cell(row=ldap_url_row, column=ldap_col).value,
            }
        }
        return dns_ntp_ldap_data

    def get_data(self):
        ipmi_data = self.get_ipmi_data()
        network_data = self.get_private_network_data()
        public_network_data = self.get_public_network_data()
        dns_ntp_ldap_data = self.get_dns_ntp_ldap_data()
        return {
            'ipmi_data': ipmi_data,
            'network_data': {
                'private': network_data,
                'public': public_network_data,
                'dns_ntp_ldap': dns_ntp_ldap_data,
            }
        }
